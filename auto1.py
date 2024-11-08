# -*- coding: utf-8 -*-
"""
Created on Thu May 13 22:10:13 2021

@author: ASUS
"""



#import openpyxl as xl
#from openpyxl.chart import LineChart, Reference

#import win32com.client
#import PIL
#from PIL import ImageGrab, Image
#import os
#import sys

import inspect

#script_path = os.path.abspath(inspect.getfile(inspect.currentframe()))
#script_name = os.path.splitext(os.path.basename(script_path))[0]
#script_dir = os.path.dirname(script_path)

#sys.path.append(script_dir + "\Modules")
#try:
#    import openpyxl
#finally:
#    del sys.path[-1]

import openpyxl as xl
from openpyxl.chart import LineChart, Reference

import win32com.client
import PIL
from PIL import ImageGrab, Image
import os
import sys

from docx.shared import Cm
from docxtpl import DocxTemplate, InlineImage
from docx.shared import Cm, Inches, Mm, Emu
import random
import datetime
import matplotlib.pyplot as plt


######## Generate automated excel workbook ########

path='D:/NONFORMAL/PYTHON/EXCEL AUTOMATION/Book1.xlsx'
workbook = xl.load_workbook(path)
sheet_1 = workbook['Sheet1']
  
for row in range(2, sheet_1.max_row + 1):
    current = sheet_1.cell(row, 2)
    voltage = sheet_1.cell(row, 3)
    power = float(current.value) * float(voltage.value)
    power_cell = sheet_1.cell(row, 1)
    power_cell.value = power
  
values = Reference(sheet_1, min_row = 2, max_row = sheet_1.max_row, min_col = 1, max_col = 1)
chart = LineChart()
chart.y_axis.title = 'Power'
chart.x_axis.title = 'Index'
chart.add_data(values)
sheet_1.add_chart(chart, 'e2')
  
workbook.save('D:/NONFORMAL/PYTHON/EXCEL AUTOMATION/Book1.xlsx')


######## Extract chart image from Excel workbook ########
path_output='D:/NONFORMAL/PYTHON/EXCEL AUTOMATION/chart.png'
input_file = path
output_image = path_output

operation = win32com.client.Dispatch("Excel.Application")
operation.Visible = 0
operation.DisplayAlerts = 0
    
workbook_2 = operation.Workbooks.Open(input_file)
sheet_2 = operation.Sheets(1)
    
for x, chart in enumerate(sheet_2.Shapes):
    chart.Copy()
    image = ImageGrab.grabclipboard()
    image.save(output_image, 'png')
    pass

workbook_2.Close(True)
operation.Quit()


######## Generating automated word document ########
template = DocxTemplate('D:/NONFORMAL/PYTHON/EXCEL AUTOMATION/template.docx')

#Generate list of random values
table_contents = []

for i in range(2, sheet_1.max_row + 1):
    
    table_contents.append({
        'Index': i-1,
        'Power': sheet_1.cell(i, 1).value,
        'Current': sheet_1.cell(i, 2).value,
        'Voltage': sheet_1.cell(i, 3).value
        })

#Import saved figure
image = InlineImage(template,'D:/NONFORMAL/PYTHON/EXCEL AUTOMATION/chart.png',Cm(10))

#Declare template variables
context = {
    'title': 'Automated Report',
    'day': datetime.datetime.now().strftime('%d'),
    'month': datetime.datetime.now().strftime('%b'),
    'year': datetime.datetime.now().strftime('%Y'),
    'table_contents': table_contents,
    'image': image
    }

#Render automated report
template.render(context)
template.save('D:/NONFORMAL/PYTHON/EXCEL AUTOMATION/Automated_report.docx')



